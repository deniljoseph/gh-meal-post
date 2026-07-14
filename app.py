#!/usr/bin/env python3
"""GeometryHome Workforce & Meal Management System v5.2 - Production Ready"""
import sqlite3,json,os,hashlib,io,base64,hmac as hmac_mod,random,difflib,asyncio,smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime,date,timedelta
from typing import Optional,List
from fastapi import FastAPI,HTTPException,Depends,UploadFile,File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse,StreamingResponse,HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer,OAuth2PasswordRequestForm
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment
import uvicorn,pandas as pd

# ── ENV CONFIG ────────────────────────────────────────────────────────────────
DATABASE_URL = os.environ.get('DATABASE_URL','')
USE_PG = DATABASE_URL.startswith(('postgres://','postgresql://'))
PORT = int(os.environ.get('PORT', 8000))
SECRET_KEY = os.environ.get('SECRET_KEY','wfms-ghome-v52-secret-change-in-prod')
APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR,'workforce.db')

oauth2_scheme = OAuth2PasswordBearer(tokenUrl='/api/auth/login',auto_error=False)

DEFAULT_RULES={
    'shift1':{'bk':True,'ln':'factory','dn':'accommodation'},
    'shift2':{'bk':True,'ln':'accommodation','dn':'factory'},
    'shift3':{'bk':True,'ln':'accommodation','dn':'accommodation'},
    'normal':{'bk':True,'ln':'factory','dn':'accommodation'},
    'site':  {'bk':True,'ln':'site','dn':'accommodation'},
    'absent_ln':'accommodation','absent_dn':'accommodation',
    'arabic_bk':False,'arabic_dn':False,
    'fasting_bk':True,'fasting_dn':True,
    'count_bk':True,'count_ln_acc':True,'count_ln_factory':True,
    'count_ln_site':True,'count_dn_acc':True,'count_dn_factory':True,'count_iftar':True,
}

# ── DATABASE LAYER ─────────────────────────────────────────────────────────────
# One fresh connection per request — no pool, no stale state, no lock leaks.
# Railway PostgreSQL handles ~100 concurrent connections; this app never
# needs more than a handful at once.
def db_conn():
    if USE_PG:
        import psycopg2
        url = DATABASE_URL.replace('postgres://','postgresql://',1)
        conn = psycopg2.connect(url)
        conn.autocommit = False
        return conn
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    c.execute('PRAGMA foreign_keys=ON')
    return c

def _sql(s):
    return s.replace('?','%s') if USE_PG else s

def q(c,s,p=()):
    if USE_PG:
        import psycopg2.extras
        try:
            cur = c.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute(_sql(s), p if p else None)
            res = [dict(r) for r in cur.fetchall()]
            cur.close(); return res
        except Exception:
            try: c.rollback()
            except: pass
            raise
    return [dict(r) for r in c.execute(s,p).fetchall()]

def q1(c,s,p=()):
    r=q(c,s,p); return r[0] if r else None

# Tables without serial 'id' column — RETURNING id would fail on these
_NO_ID_TABLES = ('settings','meal_prices')

def run(c,s,p=()):
    if USE_PG:
        try:
            cur = c.cursor()
            sql = _sql(s)
            if sql.strip().upper().startswith('INSERT'):
                has_id = not any(t in sql.lower() for t in _NO_ID_TABLES)
                if has_id:
                    cur.execute(sql+' RETURNING id', p if p else None)
                    row = cur.fetchone(); c.commit(); cur.close()
                    return row[0] if row else None
                else:
                    cur.execute(sql, p if p else None)
                    c.commit(); cur.close(); return None
            cur.execute(sql, p if p else None)
            c.commit(); cur.close(); return None
        except Exception:
            try: c.rollback()
            except: pass
            raise
    cur = c.execute(s,p); c.commit(); return cur.lastrowid

def exe(c,s,p=(),silent=False):
    """Execute DML. Raises on error unless silent=True (used in init_db)."""
    if USE_PG:
        try:
            cur = c.cursor()
            cur.execute(_sql(s), p if p else None)
            c.commit(); cur.close()
        except Exception:
            try: c.rollback()
            except: pass
            if not silent: raise
        return
    try: c.execute(s,p); c.commit()
    except:
        if not silent: raise

def rows(r): return [dict(x) for x in r]

def upsert_setting(c, key, value):
    if USE_PG:
        run(c,'INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value',(key,value))
    else:
        run(c,'INSERT OR REPLACE INTO settings VALUES(?,?)',(key,value))

def upsert_report(c, report_date, report_data, generated_by):
    if USE_PG:
        run(c,'INSERT INTO report_history(report_date,report_data,generated_by) VALUES(?,?,?) ON CONFLICT(report_date) DO UPDATE SET report_data=EXCLUDED.report_data,generated_by=EXCLUDED.generated_by',(report_date,report_data,generated_by))
    else:
        run(c,'INSERT OR REPLACE INTO report_history(report_date,report_data,generated_by)VALUES(?,?,?)',(report_date,report_data,generated_by))

# ── INIT DB ───────────────────────────────────────────────────────────────────
def init_db():
    conn = db_conn()
    if USE_PG:
        stmts = [
            """CREATE TABLE IF NOT EXISTS users(id SERIAL PRIMARY KEY,username TEXT UNIQUE NOT NULL,password_hash TEXT NOT NULL,role TEXT NOT NULL DEFAULT 'viewer',full_name TEXT,is_active INTEGER DEFAULT 1,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS locations(id SERIAL PRIMARY KEY,name TEXT UNIQUE NOT NULL,description TEXT,loc_type TEXT DEFAULT 'accommodation',is_active INTEGER DEFAULT 1,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS food_preferences(id SERIAL PRIMARY KEY,name TEXT UNIQUE NOT NULL,category TEXT,is_active INTEGER DEFAULT 1)""",
            """CREATE TABLE IF NOT EXISTS employees(id SERIAL PRIMARY KEY,emp_id TEXT UNIQUE NOT NULL,full_name TEXT NOT NULL,department TEXT,accommodation_id INTEGER REFERENCES locations(id),shift_type TEXT DEFAULT 'normal',food_pref_id INTEGER REFERENCES food_preferences(id),no_food_sunday INTEGER DEFAULT 0,remarks TEXT,status TEXT DEFAULT 'active',created_at TEXT DEFAULT CURRENT_TIMESTAMP,updated_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS attendance(id SERIAL PRIMARY KEY,employee_id INTEGER REFERENCES employees(id),att_date TEXT NOT NULL,status TEXT DEFAULT 'absent',reason TEXT,marked_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP,UNIQUE(employee_id,att_date))""",
            """CREATE TABLE IF NOT EXISTS suspensions(id SERIAL PRIMARY KEY,employee_id INTEGER REFERENCES employees(id),start_date TEXT NOT NULL,end_date TEXT,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS fasting_records(id SERIAL PRIMARY KEY,employee_id INTEGER REFERENCES employees(id),start_date TEXT NOT NULL,end_date TEXT,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS vacation_records(id SERIAL PRIMARY KEY,employee_id INTEGER REFERENCES employees(id),start_date TEXT NOT NULL,end_date TEXT NOT NULL,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS meal_exceptions(id SERIAL PRIMARY KEY,employee_id INTEGER REFERENCES employees(id),meal_type TEXT NOT NULL,start_date TEXT NOT NULL,end_date TEXT NOT NULL,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS temp_meal_overrides(id SERIAL PRIMARY KEY,employee_id INTEGER REFERENCES employees(id),override_shift_type TEXT,override_accommodation_id INTEGER REFERENCES locations(id),orig_shift_type TEXT,orig_accommodation_id INTEGER,orig_acc_name TEXT,start_date TEXT NOT NULL,end_date TEXT NOT NULL,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS report_history(id SERIAL PRIMARY KEY,report_date TEXT NOT NULL UNIQUE,report_data TEXT NOT NULL,generated_by TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS meal_prices(food_pref TEXT NOT NULL,meal_type TEXT NOT NULL,price REAL DEFAULT 0,PRIMARY KEY(food_pref,meal_type))""",
            """CREATE TABLE IF NOT EXISTS holidays(id SERIAL PRIMARY KEY,date TEXT NOT NULL UNIQUE,name TEXT NOT NULL,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY,value TEXT)""",
            """CREATE TABLE IF NOT EXISTS audit_logs(id SERIAL PRIMARY KEY,username TEXT,action TEXT NOT NULL,entity TEXT,entity_id INTEGER,detail TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
            """CREATE TABLE IF NOT EXISTS backup_history(id SERIAL PRIMARY KEY,filename TEXT NOT NULL,data TEXT NOT NULL,size_bytes INTEGER,trigger_type TEXT DEFAULT 'scheduled',created_at TEXT DEFAULT CURRENT_TIMESTAMP)""",
        ]
        for s in stmts: exe(conn,s)
    else:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY AUTOINCREMENT,username TEXT UNIQUE NOT NULL,password_hash TEXT NOT NULL,role TEXT NOT NULL DEFAULT 'viewer',full_name TEXT,is_active INTEGER DEFAULT 1,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS locations(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE NOT NULL,description TEXT,loc_type TEXT DEFAULT 'accommodation',is_active INTEGER DEFAULT 1,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS food_preferences(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE NOT NULL,category TEXT,is_active INTEGER DEFAULT 1);
        CREATE TABLE IF NOT EXISTS employees(id INTEGER PRIMARY KEY AUTOINCREMENT,emp_id TEXT UNIQUE NOT NULL,full_name TEXT NOT NULL,department TEXT,accommodation_id INTEGER REFERENCES locations(id),shift_type TEXT DEFAULT 'normal',food_pref_id INTEGER REFERENCES food_preferences(id),no_food_sunday INTEGER DEFAULT 0,remarks TEXT,status TEXT DEFAULT 'active',created_at TEXT DEFAULT CURRENT_TIMESTAMP,updated_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS attendance(id INTEGER PRIMARY KEY AUTOINCREMENT,employee_id INTEGER REFERENCES employees(id),att_date TEXT NOT NULL,status TEXT DEFAULT 'absent',reason TEXT,marked_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP,UNIQUE(employee_id,att_date));
        CREATE TABLE IF NOT EXISTS suspensions(id INTEGER PRIMARY KEY AUTOINCREMENT,employee_id INTEGER REFERENCES employees(id),start_date TEXT NOT NULL,end_date TEXT,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS fasting_records(id INTEGER PRIMARY KEY AUTOINCREMENT,employee_id INTEGER REFERENCES employees(id),start_date TEXT NOT NULL,end_date TEXT,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS vacation_records(id INTEGER PRIMARY KEY AUTOINCREMENT,employee_id INTEGER REFERENCES employees(id),start_date TEXT NOT NULL,end_date TEXT NOT NULL,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS meal_exceptions(id INTEGER PRIMARY KEY AUTOINCREMENT,employee_id INTEGER REFERENCES employees(id),meal_type TEXT NOT NULL,start_date TEXT NOT NULL,end_date TEXT NOT NULL,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS temp_meal_overrides(id INTEGER PRIMARY KEY AUTOINCREMENT,employee_id INTEGER REFERENCES employees(id),override_shift_type TEXT,override_accommodation_id INTEGER REFERENCES locations(id),orig_shift_type TEXT,orig_accommodation_id INTEGER,orig_acc_name TEXT,start_date TEXT NOT NULL,end_date TEXT NOT NULL,reason TEXT,is_active INTEGER DEFAULT 1,created_by INTEGER,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS report_history(id INTEGER PRIMARY KEY AUTOINCREMENT,report_date TEXT NOT NULL UNIQUE,report_data TEXT NOT NULL,generated_by TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS meal_prices(food_pref TEXT NOT NULL,meal_type TEXT NOT NULL,price REAL DEFAULT 0,PRIMARY KEY(food_pref,meal_type));
        CREATE TABLE IF NOT EXISTS holidays(id INTEGER PRIMARY KEY AUTOINCREMENT,date TEXT NOT NULL UNIQUE,name TEXT NOT NULL,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY,value TEXT);
        CREATE TABLE IF NOT EXISTS audit_logs(id INTEGER PRIMARY KEY AUTOINCREMENT,username TEXT,action TEXT NOT NULL,entity TEXT,entity_id INTEGER,detail TEXT,created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS backup_history(id INTEGER PRIMARY KEY AUTOINCREMENT,filename TEXT NOT NULL,data TEXT NOT NULL,size_bytes INTEGER,trigger_type TEXT DEFAULT 'scheduled',created_at TEXT DEFAULT CURRENT_TIMESTAMP);
        """)
        for col in ['ALTER TABLE employees ADD COLUMN accommodation_id INTEGER','ALTER TABLE employees ADD COLUMN shift_type TEXT DEFAULT \'normal\'','ALTER TABLE employees ADD COLUMN no_food_sunday INTEGER DEFAULT 0','ALTER TABLE locations ADD COLUMN loc_type TEXT DEFAULT \'accommodation\'']:
            try: conn.execute(col); conn.commit()
            except: pass

    h = lambda p: hashlib.sha256(p.encode()).hexdigest()
    if not q1(conn,'SELECT value FROM settings WHERE key=?',('setup_done',)):
        if not q1(conn,'SELECT COUNT(*) c FROM users',())or q1(conn,'SELECT COUNT(*) c FROM users',())['c']==0:
            exe(conn,'INSERT INTO users(username,password_hash,role,full_name)VALUES(?,?,?,?) ON CONFLICT DO NOTHING',('admin',h('admin123'),'admin','System Administrator'))
    for loc in [('LV2 - B#1','Building 1','accommodation'),('LV2 - B#11','Building 11','accommodation'),('LV2 - B#20','Building 20','accommodation'),('LV2 - B#21','Building 21','accommodation'),('Factory','Main Factory','factory')]:
        exe(conn,'INSERT INTO locations(name,description,loc_type)VALUES(?,?,?) ON CONFLICT DO NOTHING',loc)
    exe(conn,'UPDATE food_preferences SET is_active=0',silent=True)
    for fp in [('Arabic','Arabic'),('North Indian','Indian'),('North Indian Veg','Indian'),('South Indian','Indian'),('South Indian Veg','Indian')]:
        exe(conn,'INSERT INTO food_preferences(name,category)VALUES(?,?) ON CONFLICT DO NOTHING',fp)
        exe(conn,'UPDATE food_preferences SET is_active=1 WHERE name=?',(fp[0],),silent=True)
    n = q1(conn,'SELECT COUNT(*) c FROM employees',()) or {}
    if not n.get('c',0):
        def lid(nm): r=q1(conn,'SELECT id FROM locations WHERE name=?',(nm,));return r['id'] if r else None
        def fid(nm): r=q1(conn,'SELECT id FROM food_preferences WHERE name=?',(nm,));return r['id'] if r else None
        for i,s2 in enumerate([('Ahmed Al Rashid','LV2 - B#1','shift1','Arabic'),('Kumar Selvam','LV2 - B#1','shift1','South Indian'),('Ravi Shankar','LV2 - B#21','normal','South Indian'),('Mohammed Khalid','LV2 - B#21','shift2','Arabic'),('Priya Nair','LV2 - B#1','site','South Indian'),('John Smith','LV2 - B#11','shift1','North Indian'),('Ali Hassan','LV2 - B#21','shift2','Arabic'),('Raj Kumar','LV2 - B#1','site','North Indian')],1):
            exe(conn,'INSERT INTO employees(emp_id,full_name,accommodation_id,shift_type,food_pref_id,status)VALUES(?,?,?,?,?,\'active\') ON CONFLICT DO NOTHING',(f'EMP{i:04d}',s2[0],lid(s2[1]),s2[2],fid(s2[3])))
    upsert_setting(conn,'lookup_enabled','1')
    if not q1(conn,'SELECT value FROM settings WHERE key=?',('meal_rules',)):
        upsert_setting(conn,'meal_rules','{}')
    # ── Reset all serial sequences to avoid duplicate key errors after migration ──
    if USE_PG:
        serial_tables=['users','locations','food_preferences','employees',
                       'attendance','suspensions','fasting_records','vacation_records',
                       'meal_exceptions','temp_meal_overrides','report_history','holidays']
        for tbl in serial_tables:
            try:
                cur=conn.cursor()
                cur.execute(f"SELECT setval(pg_get_serial_sequence('{tbl}','id'),COALESCE(MAX(id),0)+1,false) FROM {tbl}")
                conn.commit(); cur.close()
            except Exception: 
                try: conn.rollback()
                except: pass
    conn.close()

def log_audit(conn,username,action,entity=None,entity_id=None,detail=None):
    """Best-effort audit log write — never breaks the calling request."""
    try:
        exe(conn,'INSERT INTO audit_logs(username,action,entity,entity_id,detail)VALUES(?,?,?,?,?)',(username,action,entity,entity_id,detail),silent=True)
    except Exception:
        pass

def is_holiday(conn,d):
    r=q1(conn,'SELECT id FROM holidays WHERE date=?',(d,)); return bool(r)

def get_rules(conn):
    r=q1(conn,'SELECT value FROM settings WHERE key=?',('meal_rules',))
    if r and r['value']:
        try:
            saved=json.loads(r['value']); merged={**DEFAULT_RULES}
            for k,v in saved.items():
                if isinstance(v,dict) and isinstance(merged.get(k),dict): merged[k]={**merged[k],**v}
                else: merged[k]=v
            return merged
        except: pass
    return DEFAULT_RULES

def meal_delivery(shift_type,rules,is_absent=False,is_fasting=False,food_pref=''):
    is_arabic='arabic' in (food_pref or '').lower()
    get_bk=True if not is_arabic else bool(rules.get('arabic_bk',False))
    get_dn=True if not is_arabic else bool(rules.get('arabic_dn',False))
    if is_fasting:
        return {'get_bk':bool(rules.get('fasting_bk',True)) and get_bk,'get_ln':False,'get_dn':bool(rules.get('fasting_dn',True)) and get_dn,'ln_to_acc':False,'ln_to_factory':False,'ln_to_site':False,'dn_to_factory':False,'iftar':True}
    if is_absent:
        aln=rules.get('absent_ln','accommodation'); adn=rules.get('absent_dn','accommodation')
        return {'get_bk':get_bk,'get_ln':True,'get_dn':get_dn,'ln_to_acc':aln=='accommodation','ln_to_factory':aln=='factory','ln_to_site':False,'dn_to_factory':adn=='factory','iftar':False}
    sr=rules.get(shift_type,rules.get('normal',{}))
    ln=sr.get('ln','accommodation'); dn=sr.get('dn','accommodation')
    return {'get_bk':get_bk and bool(sr.get('bk',True)),'get_ln':True,'get_dn':get_dn,'ln_to_acc':ln=='accommodation','ln_to_factory':ln=='factory','ln_to_site':ln=='site','dn_to_factory':dn=='factory','iftar':False}

def get_excluded(conn,d):
    susp=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM suspensions WHERE is_active=1 AND start_date<=? AND (end_date IS NULL OR end_date>=?)',(d,d)))
    vac=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM vacation_records WHERE is_active=1 AND start_date<=? AND end_date>=?',(d,d)))
    absent=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM attendance WHERE att_date=? AND status=?',(d,'absent')))
    fasting=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM fasting_records WHERE is_active=1 AND start_date<=? AND (end_date IS NULL OR end_date>=?)',(d,d)))
    bk_ex=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM meal_exceptions WHERE is_active=1 AND meal_type=? AND start_date<=? AND end_date>=?',('breakfast',d,d)))
    ln_ex=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM meal_exceptions WHERE is_active=1 AND meal_type=? AND start_date<=? AND end_date>=?',('lunch',d,d)))
    dn_ex=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM meal_exceptions WHERE is_active=1 AND meal_type=? AND start_date<=? AND end_date>=?',('dinner',d,d)))
    return susp,vac,absent,fasting,bk_ex,ln_ex,dn_ex

def get_temp_overrides(conn,d):
    ov=q(conn,'SELECT t.*,ln.name new_acc_name FROM temp_meal_overrides t LEFT JOIN locations ln ON t.override_accommodation_id=ln.id WHERE t.is_active=1 AND t.start_date<=? AND t.end_date>=?',(d,d))
    return {o['employee_id']:o for o in ov}

def build_report(conn,d,sunday_mode=False):
    rules=get_rules(conn); susp,vac,absent,fasting,bk_ex,ln_ex,dn_ex=get_excluded(conn,d)
    temp_ovr=get_temp_overrides(conn,d)
    emps=q(conn,'SELECT e.id,e.shift_type,e.no_food_sunday,fp.name food_pref,l.name acc FROM employees e LEFT JOIN food_preferences fp ON e.food_pref_id=fp.id LEFT JOIN locations l ON e.accommodation_id=l.id WHERE e.status=?',('active',))
    accs_raw=q(conn,'SELECT name FROM locations WHERE loc_type=? AND is_active=1 ORDER BY name',('accommodation',))
    accs=[a['name'] for a in accs_raw]; fps=['Arabic','North Indian','North Indian Veg','South Indian','South Indian Veg']
    def mk(): return{a:{f:0 for f in fps} for a in accs}
    l2a=mk();d2f=mk();l2s=mk();i2a=mk()
    gt_bk={f:0 for f in fps};gt_ln={f:0 for f in fps};gt_dn={f:0 for f in fps};gt_if={f:0 for f in fps}
    acc_sum={a:{fp:{'bk':0,'ln':0,'dn':0,'iftar':0} for fp in fps} for a in accs}
    fa_sum={a:{fp:{'bk':0,'iftar':0,'dn':0} for fp in fps} for a in accs}
    def add(tbl,acc,fp):
        if acc not in tbl:tbl[acc]={f:0 for f in fps}
        tbl[acc][fp]=tbl[acc].get(fp,0)+1
    for e in emps:
        if e['id'] in susp or e['id'] in vac:continue
        if sunday_mode and e['no_food_sunday']:continue
        acc=e['acc'] or 'Unknown';fp=e['food_pref'] or 'Unknown'
        if fp not in fps:continue
        is_ab=e['id'] in absent;is_fa=e['id'] in fasting
        eff_shift=e['shift_type'];eff_acc=acc
        if e['id'] in temp_ovr:
            ovr=temp_ovr[e['id']]
            if ovr.get('override_shift_type'):eff_shift=ovr['override_shift_type']
            if ovr.get('override_accommodation_id') and ovr.get('new_acc_name'):eff_acc=ovr['new_acc_name']
        m=meal_delivery(eff_shift,rules,is_ab,is_fa,fp)
        in_ac=eff_acc in acc_sum and fp in acc_sum[eff_acc]
        # Auto-create fa_sum entry for any acc/fp combination
        if eff_acc not in fa_sum: fa_sum[eff_acc]={}
        if fp not in fa_sum[eff_acc]: fa_sum[eff_acc][fp]={'bk':0,'iftar':0,'dn':0}
        if m['get_bk'] and e['id'] not in bk_ex and rules.get('count_bk',True):
            gt_bk[fp]=gt_bk.get(fp,0)+1
            if is_fa: fa_sum[eff_acc][fp]['bk']+=1
            elif in_ac: acc_sum[eff_acc][fp]['bk']+=1
        if is_fa:
            if rules.get('count_iftar',True):
                gt_if[fp]=gt_if.get(fp,0)+1; add(i2a,eff_acc,fp)
                fa_sum[eff_acc][fp]['iftar']+=1
        elif m['get_ln'] and e['id'] not in ln_ex:
            if sunday_mode:
                add(l2a,eff_acc,fp)
                if rules.get('count_ln_acc',True):
                    gt_ln[fp]=gt_ln.get(fp,0)+1
                    if in_ac: acc_sum[eff_acc][fp]['ln']+=1
            else:
                if m['ln_to_acc']:add(l2a,eff_acc,fp)
                elif m['ln_to_site']:add(l2s,eff_acc,fp)
                if (m['ln_to_acc'] and rules.get('count_ln_acc',True)) or (m['ln_to_factory'] and rules.get('count_ln_factory',True)) or (m['ln_to_site'] and rules.get('count_ln_site',True)):
                    gt_ln[fp]=gt_ln.get(fp,0)+1
                    if in_ac: acc_sum[eff_acc][fp]['ln']+=1
        if m['get_dn'] and e['id'] not in dn_ex:
            if sunday_mode:
                if rules.get('count_dn_acc',True):
                    gt_dn[fp]=gt_dn.get(fp,0)+1
                    if is_fa: fa_sum[eff_acc][fp]['dn']+=1
                    elif in_ac: acc_sum[eff_acc][fp]['dn']+=1
            else:
                if m['dn_to_factory']:add(d2f,eff_acc,fp)
                if (not m['dn_to_factory'] and rules.get('count_dn_acc',True)) or (m['dn_to_factory'] and rules.get('count_dn_factory',True)):
                    gt_dn[fp]=gt_dn.get(fp,0)+1
                    if is_fa: fa_sum[eff_acc][fp]['dn']+=1
                    elif in_ac: acc_sum[eff_acc][fp]['dn']+=1
    return{'lunch_to_accommodation':l2a,'dinner_to_factory':d2f,'lunch_to_site':l2s,'iftar_to_accommodation':i2a,'accommodations':accs,'food_prefs':fps,'acc_summary':acc_sum,'fasting_summary':fa_sum,'grand_total':{'breakfast':gt_bk,'lunch':gt_ln,'dinner':gt_dn,'iftar_kit':gt_if},'is_sunday_mode':sunday_mode}

# ── AUTH ───────────────────────────────────────────────────────────────────────
def make_token(u,r):
    p=json.dumps({'sub':u,'role':r,'exp':(datetime.utcnow()+timedelta(hours=10)).isoformat()})
    sig=hmac_mod.new(SECRET_KEY.encode(),p.encode(),hashlib.sha256).hexdigest()
    return base64.b64encode(f'{p}||{sig}'.encode()).decode()

def decode_token(tok):
    try:
        d=base64.b64decode(tok.encode()).decode(); ps,sig=d.rsplit('||',1)
        if hmac_mod.new(SECRET_KEY.encode(),ps.encode(),hashlib.sha256).hexdigest()!=sig: return None
        p=json.loads(ps)
        if datetime.fromisoformat(p['exp'])<datetime.utcnow(): return None
        return p
    except: return None

def get_user(tok:str=Depends(oauth2_scheme)):
    if not tok: raise HTTPException(401,'Not authenticated')
    p=decode_token(tok)
    if not p: raise HTTPException(401,'Invalid token')
    return p

def require(*roles):
    def dep(u=Depends(get_user)):
        if u['role'] not in roles: raise HTTPException(403,'Insufficient permissions')
        return u
    return dep

# ── MODELS ─────────────────────────────────────────────────────────────────────
class EmpIn(BaseModel):
    emp_id:Optional[str]=None;full_name:str;department:Optional[str]=None
    accommodation_id:Optional[int]=None;shift_type:Optional[str]='normal'
    food_pref_id:Optional[int]=None;no_food_sunday:Optional[int]=0;remarks:Optional[str]=None
    force_duplicate:Optional[bool]=False
class BulkDel(BaseModel): ids:List[int]
class BulkUpdate(BaseModel):
    ids:List[int];shift_type:Optional[str]=None;food_pref_id:Optional[int]=None;accommodation_id:Optional[int]=None
class SuspIn(BaseModel): employee_id:int;start_date:str;end_date:Optional[str]=None;reason:Optional[str]=None
class FastIn(BaseModel): employee_id:int;start_date:str;end_date:Optional[str]=None;reason:Optional[str]=None
class VacIn(BaseModel): employee_id:int;start_date:str;end_date:str;reason:Optional[str]=None
class MexIn(BaseModel): employee_id:int;meal_type:str;start_date:str;end_date:str;reason:Optional[str]=None
class AttIn(BaseModel): employee_ids:List[int];att_date:str;status:str='absent';reason:Optional[str]=None
class LocIn(BaseModel): name:str;description:Optional[str]=None;loc_type:Optional[str]='accommodation'
class FpIn(BaseModel): name:str;category:Optional[str]=None
class SetupIn(BaseModel): full_name:str;username:str;password:str;company_name:Optional[str]='GeometryHome'
class TempOvrIn(BaseModel):
    employee_id:int;override_shift_type:Optional[str]=None;override_accommodation_id:Optional[int]=None
    orig_shift_type:Optional[str]=None;orig_accommodation_id:Optional[int]=None;orig_acc_name:Optional[str]=None
    start_date:str;end_date:str;reason:Optional[str]=None
class HolidayIn(BaseModel): date:str;name:str

# ── APP ────────────────────────────────────────────────────────────────────────
app=FastAPI(title='GeometryHome WFMS v5.2')
app.add_middleware(CORSMiddleware,allow_origins=['*'],allow_methods=['*'],allow_headers=['*'])

@app.on_event('startup')
async def _start_background_tasks():
    init_db()
    asyncio.create_task(backup_scheduler_loop())

@app.get('/api/setup/status')
def setup_status():
    conn=db_conn();s=q1(conn,'SELECT value FROM settings WHERE key=?',('setup_done',));cn=q1(conn,'SELECT value FROM settings WHERE key=?',('company_name',));conn.close()
    return{'setup_done':bool(s),'company_name':cn['value'] if cn else 'GeometryHome'}

@app.post('/api/setup')
def do_setup(data:SetupIn):
    conn=db_conn()
    if q1(conn,'SELECT value FROM settings WHERE key=?',('setup_done',)): conn.close(); raise HTTPException(400,'Done')
    h=hashlib.sha256(data.password.encode()).hexdigest()
    exe(conn,'DELETE FROM users')
    run(conn,'INSERT INTO users(username,password_hash,role,full_name)VALUES(?,?,?,?)',(data.username,h,'admin',data.full_name))
    upsert_setting(conn,'setup_done','1'); upsert_setting(conn,'company_name',data.company_name)
    conn.close(); return{'ok':True}

@app.post('/api/auth/login')
def login(form:OAuth2PasswordRequestForm=Depends()):
    conn=db_conn(); h=hashlib.sha256(form.password.encode()).hexdigest()
    u=q1(conn,'SELECT * FROM users WHERE username=? AND password_hash=? AND is_active=1',(form.username,h)); conn.close()
    if not u: raise HTTPException(401,'Invalid credentials')
    return{'access_token':make_token(u['username'],u['role']),'token_type':'bearer','role':u['role'],'full_name':u['full_name']}

@app.get('/api/meal-rules')
def get_meal_rules(_=Depends(get_user)):
    conn=db_conn(); r=get_rules(conn); conn.close(); return r

@app.post('/api/meal-rules')
def save_meal_rules(data:dict,_=Depends(require('admin','hr'))):
    conn=db_conn(); curr=get_rules(conn)
    for k,v in data.items():
        if isinstance(v,dict) and isinstance(curr.get(k),dict): curr[k]={**curr[k],**v}
        else: curr[k]=v
    upsert_setting(conn,'meal_rules',json.dumps(curr)); conn.close(); return{'ok':True}

@app.get('/api/holidays')
def list_holidays(_=Depends(get_user)):
    conn=db_conn(); r=q(conn,'SELECT * FROM holidays ORDER BY date DESC'); conn.close(); return r

@app.post('/api/holidays')
def add_holiday(data:HolidayIn,_=Depends(require('admin','hr'))):
    conn=db_conn()
    try: hid=run(conn,'INSERT INTO holidays(date,name)VALUES(?,?)',(data.date,data.name)); conn.close(); return{'id':hid}
    except: conn.close(); raise HTTPException(400,'Holiday already exists for this date')

@app.delete('/api/holidays/{hid}')
def del_holiday(hid:int,_=Depends(require('admin','hr'))):
    conn=db_conn(); exe(conn,'DELETE FROM holidays WHERE id=?',(hid,)); conn.close(); return{'ok':True}

@app.get('/api/dashboard')
def dashboard(_=Depends(get_user)):
    conn=db_conn(); d=date.today().isoformat()
    total=q1(conn,'SELECT COUNT(*) c FROM employees',())['c']
    susp_ids=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM suspensions WHERE is_active=1 AND start_date<=? AND (end_date IS NULL OR end_date>=?)',(d,d)))
    absent_ids=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM attendance WHERE att_date=? AND status=?',(d,'absent')))
    fasting_ids=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM fasting_records WHERE is_active=1 AND start_date<=? AND (end_date IS NULL OR end_date>=?)',(d,d)))
    vac_ids=set(r['employee_id'] for r in q(conn,'SELECT employee_id FROM vacation_records WHERE is_active=1 AND start_date<=? AND end_date>=?',(d,d)))
    by_loc=q(conn,'SELECT l.name,COUNT(e.id) cnt FROM employees e LEFT JOIN locations l ON e.accommodation_id=l.id GROUP BY l.name')
    by_fp=q(conn,'SELECT fp.name,COUNT(e.id) cnt FROM employees e LEFT JOIN food_preferences fp ON e.food_pref_id=fp.id GROUP BY fp.name')
    by_sh=q(conn,'SELECT shift_type,COUNT(*) cnt FROM employees GROUP BY shift_type')
    conn.close()
    return{'total_employees':total,'active_meal_count':total-len(susp_ids|vac_ids),'suspended':len(susp_ids),'absent_today':len(absent_ids),'fasting':len(fasting_ids),'on_vacation':len(vac_ids),'by_location':by_loc,'by_food':by_fp,'by_shift':by_sh}

@app.get('/api/dashboard/trends')
def trends(_=Depends(get_user)):
    conn=db_conn(); days=[(date.today()-timedelta(days=i)).isoformat() for i in range(6,-1,-1)]
    result=[{'date':d,'absent':q1(conn,'SELECT COUNT(*) c FROM attendance WHERE att_date=? AND status=?',(d,'absent'))['c']} for d in days]
    conn.close(); return result

@app.get('/api/dashboard/detail/{stat_type}')
def dashboard_detail(stat_type:str,_=Depends(get_user)):
    conn=db_conn()
    try:
        today=date.today().isoformat()
        if stat_type=='active':
            rows_=q(conn,"SELECT e.full_name,e.emp_id,l.name accommodation,fp.name food_pref,e.shift_type FROM employees e LEFT JOIN locations l ON e.accommodation_id=l.id LEFT JOIN food_preferences fp ON e.food_pref_id=fp.id ORDER BY e.full_name")
            return{'title':'All Employees','rows':rows_,'columns':['full_name','emp_id','accommodation','food_pref','shift_type']}
        elif stat_type=='meal_count':
            susp,vac,absent,fasting,_,_,_=get_excluded(conn,today)
            excluded_ids=susp|vac
            rows_=q(conn,"SELECT e.full_name,e.emp_id,l.name accommodation,fp.name food_pref FROM employees e LEFT JOIN locations l ON e.accommodation_id=l.id LEFT JOIN food_preferences fp ON e.food_pref_id=fp.id WHERE e.status='active' ORDER BY e.full_name")
            included=[r for r in rows_ if True]  # meal_count = active - (susp+vac); show all active with a flag
            for r in included:
                pass
            filtered=q(conn,"SELECT e.id,e.full_name,e.emp_id,l.name accommodation,fp.name food_pref FROM employees e LEFT JOIN locations l ON e.accommodation_id=l.id LEFT JOIN food_preferences fp ON e.food_pref_id=fp.id WHERE e.status='active' ORDER BY e.full_name")
            counted=[r for r in filtered if r['id'] not in excluded_ids]
            for r in counted: r.pop('id',None)
            return{'title':"Employees Counted in Today's Meals",'rows':counted,'columns':['full_name','emp_id','accommodation','food_pref']}
        elif stat_type=='absent':
            rows_=q(conn,"SELECT e.full_name,e.emp_id,a.reason FROM attendance a JOIN employees e ON a.employee_id=e.id WHERE a.att_date=? AND a.status='absent' ORDER BY e.full_name",(today,))
            return{'title':'Absent Today','rows':rows_,'columns':['full_name','emp_id','reason']}
        elif stat_type=='suspended':
            rows_=q(conn,"SELECT e.full_name,e.emp_id,s.start_date,s.end_date,s.reason FROM suspensions s JOIN employees e ON s.employee_id=e.id WHERE s.is_active=1 AND s.start_date<=? AND (s.end_date IS NULL OR s.end_date>=?) ORDER BY e.full_name",(today,today))
            return{'title':'Currently Suspended','rows':rows_,'columns':['full_name','emp_id','start_date','end_date','reason']}
        elif stat_type=='fasting':
            rows_=q(conn,"SELECT e.full_name,e.emp_id,f.start_date,f.end_date FROM fasting_records f JOIN employees e ON f.employee_id=e.id WHERE f.is_active=1 AND f.start_date<=? AND (f.end_date IS NULL OR f.end_date>=?) ORDER BY e.full_name",(today,today))
            return{'title':'Currently Fasting','rows':rows_,'columns':['full_name','emp_id','start_date','end_date']}
        elif stat_type=='vacation':
            rows_=q(conn,"SELECT e.full_name,e.emp_id,v.start_date,v.end_date,v.reason FROM vacation_records v JOIN employees e ON v.employee_id=e.id WHERE v.is_active=1 AND v.start_date<=? AND v.end_date>=? ORDER BY e.full_name",(today,today))
            return{'title':'Currently On Leave','rows':rows_,'columns':['full_name','emp_id','start_date','end_date','reason']}
        else:
            raise HTTPException(404,'Unknown stat type')
    finally:
        conn.close()

def compute_food_count(d:str):
    conn=db_conn();rules=get_rules(conn);susp,vac,absent,fasting,bk_ex,ln_ex,dn_ex=get_excluded(conn,d)
    emps=q(conn,'SELECT e.id,e.shift_type,e.no_food_sunday,fp.name food_pref FROM employees e LEFT JOIN food_preferences fp ON e.food_pref_id=fp.id WHERE e.status=?',('active',))
    temp_ovr=get_temp_overrides(conn,d)
    is_sun=date.fromisoformat(d).weekday()==6 or is_holiday(conn,d);conn.close()
    bk={};ln={};dn={};iftar={};bkt=lnt=dnt=iftt=0
    excl_susp=len(susp);excl_ab=len(absent);excl_fa=len(fasting);excl_vac=len(vac)
    for e in emps:
        if e['id'] in susp or e['id'] in vac:continue
        if is_sun and e['no_food_sunday']:continue
        fp=e['food_pref'] or 'Unknown';is_ab=e['id'] in absent;is_fa=e['id'] in fasting
        eff_shift=e['shift_type']
        if e['id'] in temp_ovr and temp_ovr[e['id']].get('override_shift_type'):eff_shift=temp_ovr[e['id']]['override_shift_type']
        m=meal_delivery(eff_shift,rules,is_ab,is_fa,fp)
        if m['get_bk'] and e['id'] not in bk_ex and rules.get('count_bk',True): bk[fp]=bk.get(fp,0)+1;bkt+=1
        if is_fa and rules.get('count_iftar',True): iftar[fp]=iftar.get(fp,0)+1;iftt+=1
        elif m['get_ln'] and e['id'] not in ln_ex:
            if is_sun:
                if rules.get('count_ln_acc',True): ln[fp]=ln.get(fp,0)+1;lnt+=1
            elif (m['ln_to_acc'] and rules.get('count_ln_acc',True)) or (m['ln_to_factory'] and rules.get('count_ln_factory',True)) or (m['ln_to_site'] and rules.get('count_ln_site',True)):
                ln[fp]=ln.get(fp,0)+1;lnt+=1
        if m['get_dn'] and e['id'] not in dn_ex:
            if is_sun:
                if rules.get('count_dn_acc',True): dn[fp]=dn.get(fp,0)+1;dnt+=1
            elif (not m['dn_to_factory'] and rules.get('count_dn_acc',True)) or (m['dn_to_factory'] and rules.get('count_dn_factory',True)):
                dn[fp]=dn.get(fp,0)+1;dnt+=1
    bk['TOTAL']=bkt;ln['TOTAL']=lnt;dn['TOTAL']=dnt;iftar['TOTAL']=iftt
    return{'date':d,'breakfast':bk,'lunch':ln,'dinner':dn,'iftar_kit':iftar,'excluded_suspended':excl_susp,'excluded_absent':excl_ab,'excluded_fasting':excl_fa,'excluded_vacation':excl_vac}

@app.get('/api/food-count')
def food_count(target_date:Optional[str]=None,_=Depends(get_user)):
    d=target_date or date.today().isoformat()
    return compute_food_count(d)

@app.get('/api/supplier-report')
def supplier_report(target_date:Optional[str]=None,show_sunday:bool=False,save:bool=True,user=Depends(get_user)):
    conn=db_conn(); d=target_date or date.today().isoformat()
    cn=(q1(conn,'SELECT value FROM settings WHERE key=?',('company_name',)) or {}).get('value','GeometryHome')
    is_hol=is_holiday(conn,d); is_sun=date.fromisoformat(d).weekday()==6 or is_hol
    main=build_report(conn,d,sunday_mode=is_sun)
    result={'date':d,'company_name':cn,'is_sunday':is_sun,'is_holiday':is_hol,**main}
    if show_sunday and not is_sun:
        dt=date.fromisoformat(d); diff=(6-dt.weekday())%7 or 7
        next_sun=(dt+timedelta(days=diff)).isoformat()
        result['sunday_schedule']=build_report(conn,next_sun,sunday_mode=True); result['sunday_date']=next_sun
    if save:
        try: upsert_report(conn,d,json.dumps(result),user['sub'])
        except: pass
    conn.close(); return result

@app.get('/api/report-history')
def list_history(start_date:Optional[str]=None,end_date:Optional[str]=None,_=Depends(get_user)):
    conn=db_conn()
    if start_date and end_date:
        r=q(conn,'SELECT id,report_date,generated_by,created_at FROM report_history WHERE report_date>=? AND report_date<=? ORDER BY report_date DESC',(start_date,end_date))
    else:
        r=q(conn,'SELECT id,report_date,generated_by,created_at FROM report_history ORDER BY report_date DESC LIMIT 365')
    conn.close(); return r

@app.get('/api/report-history/{date_str}')
def get_history(date_str:str,_=Depends(get_user)):
    conn=db_conn(); r=q1(conn,'SELECT * FROM report_history WHERE report_date=?',(date_str,)); conn.close()
    if not r: raise HTTPException(404,'No report for this date')
    r['report_data']=json.loads(r['report_data']); return r

@app.delete('/api/report-history/{hid}')
def del_history(hid:int,_=Depends(require('admin'))):
    conn=db_conn(); exe(conn,'DELETE FROM report_history WHERE id=?',(hid,)); conn.close(); return{'ok':True}

@app.post('/api/report-history/bulk-delete')
def bulk_del_history(data:BulkDel,_=Depends(require('admin'))):
    conn=db_conn()
    for i in data.ids: exe(conn,'DELETE FROM report_history WHERE id=?',(i,))
    conn.close(); return{'deleted':len(data.ids)}

@app.get('/api/meal-prices')
def get_prices(_=Depends(get_user)):
    conn=db_conn(); r=q(conn,'SELECT * FROM meal_prices ORDER BY food_pref,meal_type'); conn.close(); return r

@app.post('/api/meal-prices')
def save_prices(data:dict,_=Depends(require('admin','hr'))):
    conn=db_conn()
    for p in data.get('prices',[]):
        if USE_PG:
            exe(conn,'INSERT INTO meal_prices(food_pref,meal_type,price)VALUES(?,?,?) ON CONFLICT(food_pref,meal_type) DO UPDATE SET price=EXCLUDED.price',(p['food_pref'],p['meal_type'],float(p.get('price',0))))
        else:
            exe(conn,'INSERT OR REPLACE INTO meal_prices(food_pref,meal_type,price)VALUES(?,?,?)',(p['food_pref'],p['meal_type'],float(p.get('price',0))))
    conn.close(); return{'ok':True}

@app.get('/api/billing/calculate')
def calc_billing(start_date:Optional[str]=None,end_date:Optional[str]=None,target_date:Optional[str]=None,_=Depends(get_user)):
    conn=db_conn()
    prices={(r['food_pref']+'|'+r['meal_type']):r['price'] for r in q(conn,'SELECT * FROM meal_prices')}
    totals={}
    if start_date and end_date:
        sd=date.fromisoformat(start_date);ed=date.fromisoformat(end_date);curr=sd
        while curr<=ed:
            d=curr.isoformat();is_sun=curr.weekday()==6 or is_holiday(conn,d)
            rpt=build_report(conn,d,sunday_mode=is_sun);gt=rpt['grand_total']
            for meal,fp_dict in [('breakfast',gt['breakfast']),('lunch',gt['lunch']),('dinner',gt['dinner']),('iftar_kit',gt['iftar_kit'])]:
                for fp,qty in fp_dict.items(): k=fp+'|'+meal;totals[k]=totals.get(k,0)+qty
            curr+=timedelta(days=1)
    else:
        d=target_date or date.today().isoformat();start_date=end_date=d
        is_sun=date.fromisoformat(d).weekday()==6 or is_holiday(conn,d)
        rpt=build_report(conn,d,sunday_mode=is_sun);gt=rpt['grand_total']
        for meal,fp_dict in [('breakfast',gt['breakfast']),('lunch',gt['lunch']),('dinner',gt['dinner']),('iftar_kit',gt['iftar_kit'])]:
            for fp,qty in fp_dict.items(): totals[fp+'|'+meal]=qty
    conn.close()
    items=[];subtotal=0.0
    for k,qty in totals.items():
        if qty==0:continue
        parts=k.rsplit('|',1);fp=parts[0];meal=parts[1]
        up=prices.get(k,0.0);tot=round(qty*up,2);subtotal+=tot
        items.append({'food':fp,'meal':meal,'qty':qty,'unit_price':up,'total':tot})
    subtotal=round(subtotal,2);vat=round(subtotal*0.05,2);net=round(subtotal+vat,2)
    return{'start_date':start_date,'end_date':end_date,'line_items':items,'subtotal':subtotal,'vat_5pct':vat,'net_total':net}

EJ='SELECT e.*,l.name accommodation_name,fp.name food_pref_name FROM employees e LEFT JOIN locations l ON e.accommodation_id=l.id LEFT JOIN food_preferences fp ON e.food_pref_id=fp.id'

def normalize_name(s):
    return ' '.join((s or '').lower().split())

def find_similar_employees(conn,full_name,emp_id=None,exclude_id=None,threshold=0.82):
    """Returns list of {id,full_name,emp_id,similarity,match_type} for likely duplicates."""
    all_emps=q(conn,'SELECT id,full_name,emp_id FROM employees'+(' WHERE id!=?' if exclude_id else ''),(exclude_id,) if exclude_id else ())
    matches=[]
    norm_new=normalize_name(full_name)
    for e in all_emps:
        # Exact emp_id collision (only relevant if emp_id was manually specified)
        if emp_id and e['emp_id'] and e['emp_id'].strip().upper()==emp_id.strip().upper():
            matches.append({**e,'similarity':1.0,'match_type':'emp_id'})
            continue
        # Fuzzy name similarity
        ratio=difflib.SequenceMatcher(None,norm_new,normalize_name(e['full_name'])).ratio()
        if ratio>=threshold:
            matches.append({**e,'similarity':round(ratio,2),'match_type':'name'})
    matches.sort(key=lambda m:-m['similarity'])
    return matches[:5]

def next_eid(conn):
    r=q1(conn,'SELECT emp_id FROM employees ORDER BY id DESC LIMIT 1')
    if not r: return 'EMP0001'
    try: return f"EMP{int(r['emp_id'].replace('EMP','').strip())+1:04d}"
    except: return f'EMP{random.randint(1000,9999)}'

@app.get('/api/employees')
def list_emp(search:Optional[str]=None,shift_type:Optional[str]=None,food_pref:Optional[str]=None,_=Depends(get_user)):
    conn=db_conn(); conds=['1=1'];params=[]
    lop='ILIKE' if USE_PG else 'LIKE'
    if search:conds.append(f'(e.full_name {lop} ? OR e.emp_id {lop} ? OR COALESCE(e.department,\'\') {lop} ?)');params+=[f'%{search}%']*3
    if shift_type:conds.append('e.shift_type=?');params.append(shift_type)
    if food_pref:conds.append('fp.name LIKE ?');params.append(f'%{food_pref}%')
    r=q(conn,EJ+' WHERE '+(' AND '.join(conds))+' ORDER BY e.full_name',params); conn.close(); return r

@app.get('/api/employees/{eid}')
def get_emp(eid:int,_=Depends(get_user)):
    conn=db_conn(); r=q1(conn,EJ+' WHERE e.id=?',(eid,)); conn.close()
    if not r: raise HTTPException(404)
    return r

@app.get('/api/employees/check-duplicate')
def check_duplicate(full_name:str,emp_id:Optional[str]=None,exclude_id:Optional[int]=None,_=Depends(get_user)):
    conn=db_conn()
    try:
        if not full_name or len(full_name.strip())<3: return{'matches':[]}
        matches=find_similar_employees(conn,full_name,emp_id,exclude_id)
        return{'matches':matches}
    finally:
        conn.close()

@app.post('/api/employees')
def create_emp(emp:EmpIn,user=Depends(require('admin','hr'))):
    conn=db_conn(); eid=emp.emp_id or next_eid(conn)
    try:
        if not emp.force_duplicate:
            matches=find_similar_employees(conn,emp.full_name,emp.emp_id)
            if matches:
                raise HTTPException(409,detail={'message':'Possible duplicate employee detected','matches':matches})
        nid=run(conn,'INSERT INTO employees(emp_id,full_name,department,accommodation_id,shift_type,food_pref_id,no_food_sunday,remarks,status)VALUES(?,?,?,?,?,?,?,?,?)',(eid,emp.full_name,emp.department,emp.accommodation_id,emp.shift_type,emp.food_pref_id,emp.no_food_sunday,emp.remarks,'active'))
        log_audit(conn,user['sub'],'create','employee',nid,emp.full_name)
        return{'id':nid,'emp_id':eid}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400,str(e))
    finally:
        conn.close()

@app.put('/api/employees/{eid}')
def update_emp(eid:int,emp:EmpIn,user=Depends(require('admin','hr'))):
    conn=db_conn()
    try:
        exe(conn,'UPDATE employees SET full_name=?,department=?,accommodation_id=?,shift_type=?,food_pref_id=?,no_food_sunday=?,remarks=?,updated_at=CURRENT_TIMESTAMP WHERE id=?',(emp.full_name,emp.department,emp.accommodation_id,emp.shift_type,emp.food_pref_id,emp.no_food_sunday,emp.remarks,eid))
        log_audit(conn,user['sub'],'update','employee',eid,emp.full_name)
        return{'ok':True}
    except Exception as e:
        raise HTTPException(500,f'Update failed: {e}')
    finally:
        conn.close()

@app.delete('/api/employees/{eid}')
def delete_emp(eid:int,user=Depends(require('admin','hr'))):
    conn=db_conn()
    try:
        emp_name=(q1(conn,'SELECT full_name FROM employees WHERE id=?',(eid,)) or {}).get('full_name','')
        for tbl in ['suspensions','fasting_records','vacation_records','meal_exceptions','temp_meal_overrides','attendance']:
            exe(conn,f'DELETE FROM {tbl} WHERE employee_id=?',(eid,))
        exe(conn,'DELETE FROM employees WHERE id=?',(eid,))
        log_audit(conn,user['sub'],'delete','employee',eid,emp_name)
        return{'ok':True}
    except Exception as e:
        raise HTTPException(500,f'Delete failed: {e}')
    finally:
        conn.close()

@app.post('/api/employees/bulk-delete')
def bulk_del_emp(data:BulkDel,_=Depends(require('admin','hr'))):
    conn=db_conn()
    try:
        for i in data.ids:
            for tbl in ['suspensions','fasting_records','vacation_records','meal_exceptions','temp_meal_overrides','attendance']:
                exe(conn,f'DELETE FROM {tbl} WHERE employee_id=?',(i,))
            exe(conn,'DELETE FROM employees WHERE id=?',(i,))
        return{'deleted':len(data.ids)}
    except Exception as e:
        raise HTTPException(500,f'Bulk delete failed: {e}')
    finally:
        conn.close()

@app.post('/api/employees/bulk-update')
def bulk_update_emp(data:BulkUpdate,user=Depends(require('admin','hr'))):
    conn=db_conn();parts=[];vals=[]
    if data.shift_type: parts.append('shift_type=?');vals.append(data.shift_type)
    if data.food_pref_id: parts.append('food_pref_id=?');vals.append(data.food_pref_id)
    if data.accommodation_id is not None and data.accommodation_id>0: parts.append('accommodation_id=?');vals.append(data.accommodation_id)
    if not parts: conn.close(); return{'updated':0}
    sql=f'UPDATE employees SET {",".join(parts)},updated_at=CURRENT_TIMESTAMP WHERE id=?'
    for eid in data.ids: exe(conn,sql,vals+[eid])
    log_audit(conn,user['sub'],'bulk_update','employee',None,f'{len(data.ids)} employee(s): {",".join(parts)}')
    conn.close(); return{'updated':len(data.ids)}

@app.get('/api/attendance')
def get_att(att_date:Optional[str]=None,_=Depends(get_user)):
    conn=db_conn(); d=att_date or date.today().isoformat()
    r=q(conn,'SELECT a.*,e.full_name,e.emp_id FROM attendance a JOIN employees e ON a.employee_id=e.id WHERE a.att_date=?',(d,))
    conn.close(); return{'date':d,'records':r}

@app.post('/api/attendance')
def mark_att(data:AttIn,user=Depends(require('admin','hr','supervisor'))):
    conn=db_conn()
    try:
        uid=q1(conn,'SELECT id FROM users WHERE username=?',(user['sub'],))
        for eid in data.employee_ids:
            if USE_PG:
                exe(conn,'INSERT INTO attendance(employee_id,att_date,status,reason,marked_by)VALUES(?,?,?,?,?) ON CONFLICT(employee_id,att_date) DO UPDATE SET status=EXCLUDED.status,reason=EXCLUDED.reason',(eid,data.att_date,data.status,data.reason,uid['id'] if uid else None))
            else:
                exe(conn,'INSERT OR REPLACE INTO attendance(employee_id,att_date,status,reason,marked_by)VALUES(?,?,?,?,?)',(eid,data.att_date,data.status,data.reason,uid['id'] if uid else None))
        log_audit(conn,user['sub'],'mark_'+data.status,'attendance',None,f'{len(data.employee_ids)} employee(s) on {data.att_date}')
        return{'ok':True}
    except Exception as e:
        raise HTTPException(500,f'Attendance update failed: {e}')
    finally:
        conn.close()

@app.delete('/api/attendance/{aid}')
def del_att(aid:int,_=Depends(require('admin','hr','supervisor'))):
    conn=db_conn()
    try:
        exe(conn,'DELETE FROM attendance WHERE id=?',(aid,))
        return{'ok':True}
    except Exception as e:
        raise HTTPException(500,f'Delete failed: {e}')
    finally:
        conn.close()

@app.get('/api/suspensions')
def list_susp(_=Depends(get_user)):
    conn=db_conn(); r=q(conn,'SELECT s.*,e.full_name,e.emp_id FROM suspensions s JOIN employees e ON s.employee_id=e.id ORDER BY s.created_at DESC'); conn.close(); return r

@app.post('/api/suspensions')
def add_susp(data:SuspIn,user=Depends(require('admin','hr'))):
    conn=db_conn()
    try:
        uid=q1(conn,'SELECT id FROM users WHERE username=?',(user['sub'],))
        sid=run(conn,'INSERT INTO suspensions(employee_id,start_date,end_date,reason,is_active,created_by)VALUES(?,?,?,?,1,?)',(data.employee_id,data.start_date,data.end_date,data.reason,uid['id'] if uid else None))
        exe(conn,"UPDATE employees SET status='suspended' WHERE id=?",(data.employee_id,))
        log_audit(conn,user['sub'],'suspend','employee',data.employee_id,data.reason)
        return{'id':sid}
    except Exception as e:
        raise HTTPException(500,f'Suspension failed: {e}')
    finally:
        conn.close()

@app.put('/api/suspensions/{sid}')
def update_susp(sid:int,data:SuspIn,_=Depends(require('admin','hr'))):
    conn=db_conn()
    try:
        exe(conn,'UPDATE suspensions SET employee_id=?,start_date=?,end_date=?,reason=? WHERE id=?',(data.employee_id,data.start_date,data.end_date,data.reason,sid))
        return{'ok':True}
    except Exception as e:
        raise HTTPException(500,f'Update failed: {e}')
    finally:
        conn.close()

@app.put('/api/suspensions/{sid}/lift')
def lift_susp(sid:int,_=Depends(require('admin','hr'))):
    conn=db_conn(); s=q1(conn,'SELECT employee_id FROM suspensions WHERE id=?',(sid,))
    if not s: conn.close(); raise HTTPException(404)
    exe(conn,'UPDATE suspensions SET is_active=0 WHERE id=?',(sid,))
    if q1(conn,'SELECT COUNT(*) c FROM suspensions WHERE employee_id=? AND is_active=1',(s['employee_id'],))['c']==0:
        exe(conn,"UPDATE employees SET status='active' WHERE id=? AND status='suspended'",(s['employee_id'],))
    conn.close(); return{'ok':True}

@app.delete('/api/suspensions/{sid}')
def del_susp(sid:int,_=Depends(require('admin'))):
    conn=db_conn(); s=q1(conn,'SELECT employee_id FROM suspensions WHERE id=?',(sid,))
    exe(conn,'DELETE FROM suspensions WHERE id=?',(sid,))
    if s and q1(conn,'SELECT COUNT(*) c FROM suspensions WHERE employee_id=? AND is_active=1',(s['employee_id'],))['c']==0:
        exe(conn,"UPDATE employees SET status='active' WHERE id=? AND status='suspended'",(s['employee_id'],))
    conn.close(); return{'ok':True}

@app.post('/api/suspensions/bulk-delete')
def bulk_del_susp(data:BulkDel,_=Depends(require('admin','hr'))):
    conn=db_conn()
    for i in data.ids:
        s=q1(conn,'SELECT employee_id FROM suspensions WHERE id=?',(i,))
        exe(conn,'DELETE FROM suspensions WHERE id=?',(i,))
        if s and q1(conn,'SELECT COUNT(*) c FROM suspensions WHERE employee_id=? AND is_active=1',(s['employee_id'],))['c']==0:
            exe(conn,"UPDATE employees SET status='active' WHERE id=? AND status='suspended'",(s['employee_id'],))
    conn.close(); return{'deleted':len(data.ids)}

@app.get('/api/fasting')
def list_fasting(_=Depends(get_user)):
    conn=db_conn(); r=q(conn,'SELECT f.*,e.full_name,e.emp_id FROM fasting_records f JOIN employees e ON f.employee_id=e.id ORDER BY f.created_at DESC'); conn.close(); return r

@app.post('/api/fasting')
def add_fasting(data:FastIn,user=Depends(require('admin','hr','supervisor'))):
    conn=db_conn(); uid=q1(conn,'SELECT id FROM users WHERE username=?',(user['sub'],))
    fid=run(conn,'INSERT INTO fasting_records(employee_id,start_date,end_date,reason,is_active,created_by)VALUES(?,?,?,?,1,?)',(data.employee_id,data.start_date,data.end_date,data.reason,uid['id'] if uid else None))
    conn.close(); return{'id':fid}

@app.put('/api/fasting/{fid}/end')
def end_fasting(fid:int,_=Depends(require('admin','hr','supervisor'))):
    conn=db_conn(); exe(conn,'UPDATE fasting_records SET is_active=0 WHERE id=?',(fid,)); conn.close(); return{'ok':True}

@app.delete('/api/fasting/{fid}')
def del_fasting(fid:int,_=Depends(require('admin'))):
    conn=db_conn(); exe(conn,'DELETE FROM fasting_records WHERE id=?',(fid,)); conn.close(); return{'ok':True}

@app.post('/api/fasting/bulk-delete')
def bulk_del_fast(data:BulkDel,_=Depends(require('admin','hr'))):
    conn=db_conn()
    for i in data.ids: exe(conn,'DELETE FROM fasting_records WHERE id=?',(i,))
    conn.close(); return{'deleted':len(data.ids)}

@app.get('/api/vacation')
def list_vac(_=Depends(get_user)):
    conn=db_conn(); r=q(conn,'SELECT v.*,e.full_name,e.emp_id FROM vacation_records v JOIN employees e ON v.employee_id=e.id ORDER BY v.created_at DESC'); conn.close(); return r

@app.post('/api/vacation')
def add_vac(data:VacIn,user=Depends(require('admin','hr'))):
    conn=db_conn(); uid=q1(conn,'SELECT id FROM users WHERE username=?',(user['sub'],))
    vid=run(conn,'INSERT INTO vacation_records(employee_id,start_date,end_date,reason,is_active,created_by)VALUES(?,?,?,?,1,?)',(data.employee_id,data.start_date,data.end_date,data.reason,uid['id'] if uid else None))
    conn.close(); return{'id':vid}

@app.put('/api/vacation/{vid}')
def update_vac(vid:int,data:VacIn,_=Depends(require('admin','hr'))):
    conn=db_conn(); exe(conn,'UPDATE vacation_records SET start_date=?,end_date=?,reason=? WHERE id=?',(data.start_date,data.end_date,data.reason,vid)); conn.close(); return{'ok':True}

@app.delete('/api/vacation/{vid}')
def del_vac(vid:int,_=Depends(require('admin','hr'))):
    conn=db_conn(); exe(conn,'DELETE FROM vacation_records WHERE id=?',(vid,)); conn.close(); return{'ok':True}

@app.post('/api/vacation/bulk-delete')
def bulk_del_vac(data:BulkDel,_=Depends(require('admin','hr'))):
    conn=db_conn()
    for i in data.ids: exe(conn,'DELETE FROM vacation_records WHERE id=?',(i,))
    conn.close(); return{'deleted':len(data.ids)}

@app.get('/api/meal-exceptions')
def list_mex(_=Depends(get_user)):
    conn=db_conn(); r=q(conn,'SELECT me.*,e.full_name,e.emp_id FROM meal_exceptions me JOIN employees e ON me.employee_id=e.id ORDER BY me.created_at DESC'); conn.close(); return r

@app.post('/api/meal-exceptions')
def add_mex(data:MexIn,user=Depends(require('admin','hr'))):
    conn=db_conn(); uid=q1(conn,'SELECT id FROM users WHERE username=?',(user['sub'],))
    mid=run(conn,'INSERT INTO meal_exceptions(employee_id,meal_type,start_date,end_date,reason,is_active,created_by)VALUES(?,?,?,?,?,1,?)',(data.employee_id,data.meal_type,data.start_date,data.end_date,data.reason,uid['id'] if uid else None))
    conn.close(); return{'id':mid}

@app.put('/api/meal-exceptions/{mid}')
def update_mex(mid:int,data:MexIn,_=Depends(require('admin','hr'))):
    conn=db_conn(); exe(conn,'UPDATE meal_exceptions SET meal_type=?,start_date=?,end_date=?,reason=? WHERE id=?',(data.meal_type,data.start_date,data.end_date,data.reason,mid)); conn.close(); return{'ok':True}

@app.delete('/api/meal-exceptions/{mid}')
def del_mex(mid:int,_=Depends(require('admin'))):
    conn=db_conn(); exe(conn,'DELETE FROM meal_exceptions WHERE id=?',(mid,)); conn.close(); return{'ok':True}

@app.post('/api/meal-exceptions/bulk-delete')
def bulk_del_mex(data:BulkDel,_=Depends(require('admin','hr'))):
    conn=db_conn()
    for i in data.ids: exe(conn,'DELETE FROM meal_exceptions WHERE id=?',(i,))
    conn.close(); return{'deleted':len(data.ids)}

@app.get('/api/temp-overrides')
def list_overrides(_=Depends(get_user)):
    conn=db_conn(); r=q(conn,'SELECT t.*,e.full_name,e.emp_id,ln.name new_acc_name FROM temp_meal_overrides t JOIN employees e ON t.employee_id=e.id LEFT JOIN locations ln ON t.override_accommodation_id=ln.id ORDER BY t.created_at DESC'); conn.close(); return r

@app.post('/api/temp-overrides')
def add_override(data:TempOvrIn,user=Depends(require('admin','hr'))):
    conn=db_conn(); uid=q1(conn,'SELECT id FROM users WHERE username=?',(user['sub'],))
    oid=run(conn,'INSERT INTO temp_meal_overrides(employee_id,override_shift_type,override_accommodation_id,orig_shift_type,orig_accommodation_id,orig_acc_name,start_date,end_date,reason,created_by)VALUES(?,?,?,?,?,?,?,?,?,?)',(data.employee_id,data.override_shift_type,data.override_accommodation_id,data.orig_shift_type,data.orig_accommodation_id,data.orig_acc_name,data.start_date,data.end_date,data.reason,uid['id'] if uid else None))
    conn.close(); return{'id':oid}

@app.delete('/api/temp-overrides/{oid}')
def del_override(oid:int,_=Depends(require('admin','hr'))):
    conn=db_conn(); exe(conn,'DELETE FROM temp_meal_overrides WHERE id=?',(oid,)); conn.close(); return{'ok':True}

@app.post('/api/temp-overrides/bulk-delete')
def bulk_del_overrides(data:BulkDel,_=Depends(require('admin','hr'))):
    conn=db_conn()
    for i in data.ids: exe(conn,'DELETE FROM temp_meal_overrides WHERE id=?',(i,))
    conn.close(); return{'deleted':len(data.ids)}

@app.get('/api/locations')
def list_locs(_=Depends(get_user)):
    conn=db_conn(); r=q(conn,'SELECT * FROM locations WHERE is_active=1 ORDER BY name'); conn.close(); return r

@app.post('/api/locations')
def add_loc(data:LocIn,_=Depends(require('admin','hr'))):
    conn=db_conn(); lid=run(conn,'INSERT INTO locations(name,description,loc_type)VALUES(?,?,?)',(data.name,data.description,data.loc_type)); conn.close(); return{'id':lid}

@app.put('/api/locations/{lid}')
def update_loc(lid:int,data:LocIn,_=Depends(require('admin','hr'))):
    conn=db_conn(); exe(conn,'UPDATE locations SET name=?,description=?,loc_type=? WHERE id=?',(data.name,data.description,data.loc_type,lid)); conn.close(); return{'ok':True}

@app.delete('/api/locations/{lid}')
def del_loc(lid:int,_=Depends(require('admin'))):
    conn=db_conn(); exe(conn,'UPDATE locations SET is_active=0 WHERE id=?',(lid,)); conn.close(); return{'ok':True}

@app.get('/api/food-preferences')
def list_fp(_=Depends(get_user)):
    conn=db_conn(); r=q(conn,'SELECT * FROM food_preferences WHERE is_active=1 ORDER BY name'); conn.close(); return r

@app.post('/api/food-preferences')
def add_fp(data:FpIn,_=Depends(require('admin','hr'))):
    conn=db_conn(); fid=run(conn,'INSERT INTO food_preferences(name,category)VALUES(?,?)',(data.name,data.category)); conn.close(); return{'id':fid}

@app.put('/api/food-preferences/{fid}')
def update_fp(fid:int,data:FpIn,_=Depends(require('admin','hr'))):
    conn=db_conn(); exe(conn,'UPDATE food_preferences SET name=?,category=? WHERE id=?',(data.name,data.category,fid)); conn.close(); return{'ok':True}

@app.delete('/api/food-preferences/{fid}')
def del_fp(fid:int,_=Depends(require('admin'))):
    conn=db_conn(); exe(conn,'UPDATE food_preferences SET is_active=0 WHERE id=?',(fid,)); conn.close(); return{'ok':True}

@app.get('/api/users')
def list_users(_=Depends(require('admin'))):
    conn=db_conn(); r=q(conn,'SELECT id,username,role,full_name,is_active,created_at FROM users ORDER BY full_name'); conn.close(); return r

@app.post('/api/users')
def add_user(data:dict,_=Depends(require('admin'))):
    conn=db_conn(); h=hashlib.sha256(data['password'].encode()).hexdigest()
    uid=run(conn,'INSERT INTO users(username,password_hash,role,full_name)VALUES(?,?,?,?)',(data['username'],h,data['role'],data['full_name'])); conn.close(); return{'id':uid}

@app.put('/api/users/{uid}')
def update_user(uid:int,data:dict,_=Depends(require('admin'))):
    conn=db_conn()
    if data.get('password'): h=hashlib.sha256(data['password'].encode()).hexdigest();exe(conn,'UPDATE users SET role=?,full_name=?,password_hash=? WHERE id=?',(data['role'],data['full_name'],h,uid))
    else: exe(conn,'UPDATE users SET role=?,full_name=? WHERE id=?',(data['role'],data['full_name'],uid))
    conn.close(); return{'ok':True}

@app.delete('/api/users/{uid}')
def del_user(uid:int,_=Depends(require('admin'))):
    conn=db_conn(); exe(conn,'UPDATE users SET is_active=0 WHERE id=?',(uid,)); conn.close(); return{'ok':True}

@app.post('/api/import/employees')
async def import_emp(file:UploadFile=File(...),user=Depends(require('admin','hr'))):
    content=await file.read()
    try: df=pd.read_excel(io.BytesIO(content))
    except Exception as e: raise HTTPException(400,f'Cannot read: {e}')
    df.columns=[str(c).strip().lower().replace(' ','_').replace('-','_') for c in df.columns]
    COL={'emp_id':['emp_id','employee_id','id'],'full_name':['full_name','name','employee_name'],'department':['department','dept'],'food_pref':['food_preference','food_pref','food'],'location':['location','accommodation','block','camp'],'shift':['shift','shift_type','employee_type','type']}
    def find(k):
        for a in COL.get(k,[k]):
            if a in df.columns: return a
        return None
    def g(row,k,default=None):
        col=find(k);v=row.get(col) if col else None
        return None if (v is None or str(v)=='nan') else str(v).strip()
    conn=db_conn();imported=skipped=0;errors=[]
    for _,row in df.iterrows():
        row=row.to_dict();name=g(row,'full_name')
        if not name or name.lower()=='nan': skipped+=1; continue
        eid=g(row,'emp_id') or next_eid(conn)
        def rloc(val):
            if not val: return None
            r2=q1(conn,'SELECT id FROM locations WHERE LOWER(name) LIKE ?',(f'%{val.lower()}%',))
            return r2['id'] if r2 else run(conn,'INSERT INTO locations(name,loc_type)VALUES(?,?)',(val,'accommodation'))
        def rfp(val):
            if not val: return None
            r2=q1(conn,'SELECT id FROM food_preferences WHERE LOWER(name) LIKE ? AND is_active=1',(f'%{val.lower()}%',))
            return r2['id'] if r2 else None
        def rst(val):
            if not val: return 'normal'
            v=val.lower()
            if '1' in v or 'first' in v: return 'shift1'
            if '2' in v or 'second' in v: return 'shift2'
            if '3' in v or 'third' in v: return 'shift3'
            if 'site' in v: return 'site'
            return 'normal'
        try:
            cur=run(conn,'INSERT INTO employees(emp_id,full_name,department,accommodation_id,shift_type,food_pref_id,status)VALUES(?,?,?,?,?,?,?) ON CONFLICT DO NOTHING',(eid,name,g(row,'department'),rloc(g(row,'location')),rst(g(row,'shift','normal')),rfp(g(row,'food_pref')),'active'))
            if cur: imported+=1
            else: skipped+=1
        except Exception as ex: errors.append(str(ex)[:80])
    conn.close(); return{'imported':imported,'skipped':skipped,'errors':errors[:10]}

@app.get('/api/export/employees')
def export_emp(_=Depends(get_user)):
    conn=db_conn()
    emps=q(conn,'SELECT e.emp_id,e.full_name,e.department,e.shift_type,l.name accommodation,fp.name food_preference,e.no_food_sunday,e.remarks FROM employees e LEFT JOIN locations l ON e.accommodation_id=l.id LEFT JOIN food_preferences fp ON e.food_pref_id=fp.id')
    conn.close()
    wb=openpyxl.Workbook();ws=wb.active;ws.title='Employees'
    hf=PatternFill('solid',fgColor='003c8f');hfont=Font(bold=True,color='FFFFFF')
    hdrs=['Emp ID','Full Name','Department','Shift','Accommodation','Food Preference','No Food Sunday','Remarks']
    keys=['emp_id','full_name','department','shift_type','accommodation','food_preference','no_food_sunday','remarks']
    for i,h in enumerate(hdrs,1):
        c=ws.cell(1,i,h);c.fill=hf;c.font=hfont;c.alignment=Alignment(horizontal='center')
    for ri,e in enumerate(emps,2):
        for ci,k in enumerate(keys,1): ws.cell(ri,ci,e.get(k,''))
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=18
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return StreamingResponse(buf,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':'attachment; filename=employees.xlsx'})

@app.get('/api/export/food-report')
def export_food(target_date:Optional[str]=None,_=Depends(get_user)):
    d=target_date or date.today().isoformat()
    data=compute_food_count(d)
    wb=openpyxl.Workbook();ws=wb.active;ws.title=f'Food {d}'
    hf=PatternFill('solid',fgColor='003c8f');hfont=Font(bold=True,color='FFFFFF')
    ws['A1']=f'Daily Food Report - {d}';ws['A1'].font=Font(bold=True,size=14);ws.merge_cells('A1:E1')
    for i,h in enumerate(['Meal','Food Category','Count','Date'],1):
        c=ws.cell(3,i,h);c.fill=hf;c.font=hfont
    row=4
    for meal in['breakfast','lunch','dinner','iftar_kit']:
        for cat,cnt in data[meal].items():
            if cat=='TOTAL':continue
            ws.cell(row,1,meal.capitalize());ws.cell(row,2,cat);ws.cell(row,3,cnt);ws.cell(row,4,d);row+=1
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=20
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return StreamingResponse(buf,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':f'attachment; filename=food_report_{d}.xlsx'})

# ── FULL DATABASE BACKUP / RESTORE ─────────────────────────────────────────────
# Tables in dependency order (parents before children) — used for both
# export sheet order and restore insert order.
BACKUP_TABLES=['settings','locations','food_preferences','users','employees',
    'attendance','suspensions','fasting_records','vacation_records',
    'meal_exceptions','temp_meal_overrides','report_history','meal_prices','holidays','audit_logs']

@app.get('/api/audit-log')
def get_audit_log(limit:int=200,_=Depends(require('admin'))):
    conn=db_conn()
    try:
        r=q(conn,'SELECT * FROM audit_logs ORDER BY id DESC LIMIT ?',(limit,))
        return r
    finally:
        conn.close()

class BulkSuspIn(BaseModel):
    employee_ids:List[int];start_date:str;end_date:Optional[str]=None;reason:Optional[str]=None
class BulkVacIn(BaseModel):
    employee_ids:List[int];start_date:str;end_date:str;reason:Optional[str]=None
class BulkFastIn(BaseModel):
    employee_ids:List[int];start_date:str;end_date:Optional[str]=None;reason:Optional[str]=None

@app.post('/api/suspensions/bulk-add')
def bulk_add_susp(data:BulkSuspIn,user=Depends(require('admin','hr'))):
    conn=db_conn()
    try:
        uid=q1(conn,'SELECT id FROM users WHERE username=?',(user['sub'],))
        n=0
        for eid in data.employee_ids:
            run(conn,'INSERT INTO suspensions(employee_id,start_date,end_date,reason,is_active,created_by)VALUES(?,?,?,?,1,?)',(eid,data.start_date,data.end_date,data.reason,uid['id'] if uid else None))
            exe(conn,"UPDATE employees SET status='suspended' WHERE id=?",(eid,))
            n+=1
        log_audit(conn,user['sub'],'bulk_suspend','employee',None,f'{n} employee(s)')
        return{'created':n}
    except Exception as e:
        raise HTTPException(500,f'Bulk suspend failed: {e}')
    finally:
        conn.close()

@app.post('/api/vacation/bulk-add')
def bulk_add_vac(data:BulkVacIn,user=Depends(require('admin','hr'))):
    conn=db_conn()
    try:
        uid=q1(conn,'SELECT id FROM users WHERE username=?',(user['sub'],))
        n=0
        for eid in data.employee_ids:
            run(conn,'INSERT INTO vacation_records(employee_id,start_date,end_date,reason,is_active,created_by)VALUES(?,?,?,?,1,?)',(eid,data.start_date,data.end_date,data.reason,uid['id'] if uid else None))
            n+=1
        log_audit(conn,user['sub'],'bulk_vacation','employee',None,f'{n} employee(s)')
        return{'created':n}
    except Exception as e:
        raise HTTPException(500,f'Bulk vacation failed: {e}')
    finally:
        conn.close()

@app.post('/api/fasting/bulk-add')
def bulk_add_fast(data:BulkFastIn,user=Depends(require('admin','hr','supervisor'))):
    conn=db_conn()
    try:
        uid=q1(conn,'SELECT id FROM users WHERE username=?',(user['sub'],))
        n=0
        for eid in data.employee_ids:
            run(conn,'INSERT INTO fasting_records(employee_id,start_date,end_date,reason,is_active,created_by)VALUES(?,?,?,?,1,?)',(eid,data.start_date,data.end_date,data.reason,uid['id'] if uid else None))
            n+=1
        log_audit(conn,user['sub'],'bulk_fasting','employee',None,f'{n} employee(s)')
        return{'created':n}
    except Exception as e:
        raise HTTPException(500,f'Bulk fasting failed: {e}')
    finally:
        conn.close()

@app.get('/api/alerts')
def get_alerts(_=Depends(get_user)):
    conn=db_conn()
    try:
        today=date.today().isoformat()
        soon=(date.today()+timedelta(days=3)).isoformat()
        alerts=[]
        # Suspensions expiring soon
        exp_susp=q(conn,"SELECT s.id,e.full_name,s.end_date FROM suspensions s JOIN employees e ON s.employee_id=e.id WHERE s.is_active=1 AND s.end_date IS NOT NULL AND s.end_date>=? AND s.end_date<=?",(today,soon))
        for r in exp_susp:
            alerts.append({'type':'suspension_expiring','severity':'warning','message':f"{r['full_name']}'s suspension ends {r['end_date']}",'link':'suspensions'})
        # Temp overrides expiring soon
        exp_ovr=q(conn,"SELECT t.id,e.full_name,t.end_date FROM temp_meal_overrides t JOIN employees e ON t.employee_id=e.id WHERE t.is_active=1 AND t.end_date>=? AND t.end_date<=?",(today,soon))
        for r in exp_ovr:
            alerts.append({'type':'override_expiring','severity':'info','message':f"{r['full_name']}'s temporary override ends {r['end_date']}",'link':'temp-overrides'})
        # Vacation ending soon (returning to work)
        end_vac=q(conn,"SELECT v.id,e.full_name,v.end_date FROM vacation_records v JOIN employees e ON v.employee_id=e.id WHERE v.is_active=1 AND v.end_date>=? AND v.end_date<=?",(today,soon))
        for r in end_vac:
            alerts.append({'type':'vacation_ending','severity':'info','message':f"{r['full_name']} returns from leave {r['end_date']}",'link':'vacation'})
        # Employees missing accommodation
        no_acc=q(conn,"SELECT COUNT(*) c FROM employees WHERE status='active' AND accommodation_id IS NULL")
        if no_acc and no_acc[0]['c']>0:
            alerts.append({'type':'missing_accommodation','severity':'danger','message':f"{no_acc[0]['c']} active employee(s) have no accommodation assigned",'link':'employees'})
        # Employees missing food preference
        no_fp=q(conn,"SELECT COUNT(*) c FROM employees WHERE status='active' AND food_pref_id IS NULL")
        if no_fp and no_fp[0]['c']>0:
            alerts.append({'type':'missing_food_pref','severity':'danger','message':f"{no_fp[0]['c']} active employee(s) have no food preference set",'link':'employees'})
        # No report generated today
        rpt_today=q1(conn,'SELECT id FROM report_history WHERE report_date=?',(today,))
        if not rpt_today:
            alerts.append({'type':'no_report_today','severity':'warning','message':"No supplier report generated yet today",'link':'supplier-report'})
        return{'alerts':alerts,'count':len(alerts)}
    finally:
        conn.close()

@app.get('/api/export/csv/{section}')
def export_csv(section:str,_=Depends(get_user)):
    conn=db_conn()
    try:
        queries={
            'suspensions':"SELECT e.full_name,e.emp_id,s.start_date,s.end_date,s.reason,s.is_active FROM suspensions s JOIN employees e ON s.employee_id=e.id ORDER BY s.start_date DESC",
            'vacation':"SELECT e.full_name,e.emp_id,v.start_date,v.end_date,v.reason FROM vacation_records v JOIN employees e ON v.employee_id=e.id ORDER BY v.start_date DESC",
            'fasting':"SELECT e.full_name,e.emp_id,f.start_date,f.end_date,f.reason,f.is_active FROM fasting_records f JOIN employees e ON f.employee_id=e.id ORDER BY f.start_date DESC",
            'meal-exceptions':"SELECT e.full_name,e.emp_id,m.meal_type,m.start_date,m.end_date,m.reason FROM meal_exceptions m JOIN employees e ON m.employee_id=e.id ORDER BY m.start_date DESC",
            'temp-overrides':"SELECT e.full_name,e.emp_id,t.override_shift_type,t.orig_acc_name,t.start_date,t.end_date,t.reason FROM temp_meal_overrides t JOIN employees e ON t.employee_id=e.id ORDER BY t.start_date DESC",
            'attendance':"SELECT e.full_name,e.emp_id,a.att_date,a.status,a.reason FROM attendance a JOIN employees e ON a.employee_id=e.id ORDER BY a.att_date DESC",
            'audit-log':"SELECT username,action,entity,entity_id,detail,created_at FROM audit_logs ORDER BY id DESC LIMIT 500",
        }
        if section not in queries:
            raise HTTPException(404,'Unknown export section')
        rows_=q(conn,queries[section])
        buf=io.StringIO()
        if rows_:
            import csv
            w=csv.DictWriter(buf,fieldnames=list(rows_[0].keys()))
            w.writeheader()
            for r in rows_: w.writerow(r)
        else:
            buf.write('No records\n')
        buf.seek(0)
        return StreamingResponse(io.BytesIO(buf.getvalue().encode()),media_type='text/csv',headers={'Content-Disposition':f'attachment; filename={section}_{date.today().isoformat()}.csv'})
    finally:
        conn.close()

def build_backup_xlsx(conn):
    """Builds the full backup workbook and returns raw bytes. Shared by manual download and scheduled task."""
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    hf=PatternFill('solid',fgColor='003c8f'); hfont=Font(bold=True,color='FFFFFF')
    for tbl in BACKUP_TABLES:
        rows_=q(conn,f'SELECT * FROM {tbl}')
        ws=wb.create_sheet(title=tbl[:31])  # Excel sheet name limit
        if rows_:
            cols=list(rows_[0].keys())
            for i,h in enumerate(cols,1):
                c=ws.cell(1,i,h); c.fill=hf; c.font=hfont
            for ri,row in enumerate(rows_,2):
                for ci,col in enumerate(cols,1):
                    val=row.get(col)
                    if isinstance(val,(dict,list)): val=json.dumps(val)
                    ws.cell(ri,ci,val)
            for col_cells in ws.columns:
                ws.column_dimensions[col_cells[0].column_letter].width=18
        else:
            ws.cell(1,1,'(empty table)')
    meta=wb.create_sheet(title='_backup_info',index=0)
    meta['A1']='GH Meal Management System - Full Backup'; meta['A1'].font=Font(bold=True,size=13)
    meta['A2']=f'Created: {datetime.utcnow().isoformat()} UTC'
    meta['A3']=f'Database: {"PostgreSQL" if USE_PG else "SQLite"}'
    meta['A4']=f'Tables: {len(BACKUP_TABLES)}'
    meta['A6']='⚠️  Restoring this file will REPLACE ALL current data. Use with caution.'
    meta['A6'].font=Font(bold=True,color='B71C1C')
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

@app.get('/api/backup/export')
def backup_export(_=Depends(require('admin'))):
    conn=db_conn()
    try:
        data=build_backup_xlsx(conn)
        fname=f'GH_Backup_{date.today().isoformat()}.xlsx'
        return StreamingResponse(io.BytesIO(data),media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':f'attachment; filename={fname}'})
    finally:
        conn.close()

def send_backup_email(data:bytes,fname:str):
    """Send backup as an email attachment if SMTP env vars are configured. Silent no-op otherwise."""
    smtp_host=os.environ.get('SMTP_HOST')
    smtp_user=os.environ.get('SMTP_USER')
    smtp_pass=os.environ.get('SMTP_PASS')
    to_addr=os.environ.get('BACKUP_EMAIL_TO')
    if not (smtp_host and smtp_user and smtp_pass and to_addr):
        return False
    try:
        smtp_port=int(os.environ.get('SMTP_PORT','587'))
        msg=MIMEMultipart()
        msg['From']=smtp_user; msg['To']=to_addr
        msg['Subject']=f'GH Meal System — Daily Backup {date.today().isoformat()}'
        msg.attach(MIMEText(f'Automatic daily backup attached.\n\nGenerated: {datetime.utcnow().isoformat()} UTC\nDatabase: {"PostgreSQL" if USE_PG else "SQLite"}','plain'))
        part=MIMEBase('application','octet-stream'); part.set_payload(data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',f'attachment; filename={fname}')
        msg.attach(part)
        with smtplib.SMTP(smtp_host,smtp_port,timeout=20) as server:
            server.starttls(); server.login(smtp_user,smtp_pass); server.send_message(msg)
        return True
    except Exception as ex:
        print(f'[backup email] failed: {ex}')
        return False

def run_scheduled_backup():
    """Generates a backup, stores it in backup_history, optionally emails it, and prunes old entries."""
    conn=db_conn()
    try:
        data=build_backup_xlsx(conn)
        fname=f'GH_Backup_{date.today().isoformat()}.xlsx'
        b64=base64.b64encode(data).decode()
        run(conn,'INSERT INTO backup_history(filename,data,size_bytes,trigger_type)VALUES(?,?,?,?)',(fname,b64,len(data),'scheduled'))
        # Keep only the most recent 14 scheduled backups
        old=q(conn,"SELECT id FROM backup_history WHERE trigger_type='scheduled' ORDER BY id DESC")
        for row in old[14:]:
            exe(conn,'DELETE FROM backup_history WHERE id=?',(row['id'],),silent=True)
        emailed=send_backup_email(data,fname)
        log_audit(conn,'system','scheduled_backup',None,None,f'{len(data)} bytes'+(', emailed' if emailed else ', stored only'))
        print(f'[scheduled backup] completed: {fname} ({len(data)} bytes){" — emailed" if emailed else ""}')
    except Exception as ex:
        print(f'[scheduled backup] FAILED: {ex}')
    finally:
        conn.close()

async def backup_scheduler_loop():
    """Runs once per day at ~02:00 server time. Also does an initial catch-up run 60s after startup
    if no backup exists yet for today, so restarts on Railway don't cause missed days silently."""
    await asyncio.sleep(60)  # let the app fully boot first
    try:
        conn=db_conn()
        today=date.today().isoformat()
        existing=q1(conn,"SELECT id FROM backup_history WHERE created_at LIKE ? ORDER BY id DESC LIMIT 1",(today+'%',))
        conn.close()
        if not existing:
            run_scheduled_backup()
    except Exception as ex:
        print(f'[scheduled backup] startup check failed: {ex}')
    while True:
        now=datetime.utcnow()
        next_run=(now+timedelta(days=1)).replace(hour=2,minute=0,second=0,microsecond=0)
        if next_run<=now: next_run+=timedelta(days=1)
        await asyncio.sleep(max(60,(next_run-now).total_seconds()))
        run_scheduled_backup()

@app.get('/api/backup/history')
def backup_history_list(_=Depends(require('admin'))):
    conn=db_conn()
    try:
        rows_=q(conn,'SELECT id,filename,size_bytes,trigger_type,created_at FROM backup_history ORDER BY id DESC LIMIT 30')
        return rows_
    finally:
        conn.close()

@app.get('/api/backup/download/{backup_id}')
def backup_download(backup_id:int,_=Depends(require('admin'))):
    conn=db_conn()
    try:
        r=q1(conn,'SELECT filename,data FROM backup_history WHERE id=?',(backup_id,))
        if not r: raise HTTPException(404,'Backup not found')
        data=base64.b64decode(r['data'])
        return StreamingResponse(io.BytesIO(data),media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':f'attachment; filename={r["filename"]}'})
    finally:
        conn.close()

@app.delete('/api/backup/history/{backup_id}')
def backup_history_delete(backup_id:int,_=Depends(require('admin'))):
    conn=db_conn()
    try:
        exe(conn,'DELETE FROM backup_history WHERE id=?',(backup_id,))
        return{'ok':True}
    finally:
        conn.close()

@app.post('/api/backup/run-now')
def backup_run_now(user=Depends(require('admin'))):
    conn=db_conn()
    try:
        data=build_backup_xlsx(conn)
        fname=f'GH_Backup_{date.today().isoformat()}_manual.xlsx'
        b64=base64.b64encode(data).decode()
        bid=run(conn,'INSERT INTO backup_history(filename,data,size_bytes,trigger_type)VALUES(?,?,?,?)',(fname,b64,len(data),'manual'))
        emailed=send_backup_email(data,fname)
        log_audit(conn,user['sub'],'manual_backup',None,bid,f'{len(data)} bytes'+(', emailed' if emailed else ''))
        return{'id':bid,'filename':fname,'size_bytes':len(data),'emailed':emailed}
    finally:
        conn.close()

@app.post('/api/backup/restore')
async def backup_restore(file:UploadFile=File(...),user=Depends(require('admin'))):
    content=await file.read()
    try:
        wb=openpyxl.load_workbook(io.BytesIO(content),data_only=True)
    except Exception as e:
        raise HTTPException(400,f'Cannot read backup file: {e}')

    conn=db_conn()
    try:
        restored={}
        # Wipe in reverse dependency order (children first) to satisfy FKs
        for tbl in reversed(BACKUP_TABLES):
            exe(conn,f'DELETE FROM {tbl}',silent=True)

        for tbl in BACKUP_TABLES:
            if tbl not in wb.sheetnames:
                restored[tbl]=0; continue
            ws=wb[tbl]
            rows_iter=list(ws.iter_rows(values_only=True))
            if not rows_iter or len(rows_iter)<2:
                restored[tbl]=0; continue
            headers=[str(h) for h in rows_iter[0] if h is not None]
            count=0
            for r in rows_iter[1:]:
                if all(v is None for v in r): continue
                vals=list(r[:len(headers)])
                placeholders=','.join(['?']*len(headers))
                col_str=','.join(headers)
                try:
                    exe(conn,f'INSERT INTO {tbl}({col_str}) VALUES({placeholders})',vals)
                    count+=1
                except Exception:
                    pass  # skip malformed row, keep going
            restored[tbl]=count

        # Reset PG sequences after restore so new inserts don't collide
        if USE_PG:
            for tbl in BACKUP_TABLES:
                if tbl in ('settings','meal_prices'): continue
                try:
                    cur=conn.cursor()
                    cur.execute(f"SELECT setval(pg_get_serial_sequence('{tbl}','id'),COALESCE(MAX(id),0)+1,false) FROM {tbl}")
                    conn.commit(); cur.close()
                except Exception:
                    try: conn.rollback()
                    except: pass

        log_audit(conn,user['sub'],'restore','backup',None,f'{sum(restored.values())} rows restored from {file.filename}')
        return{'ok':True,'restored':restored,'total':sum(restored.values())}
    except Exception as e:
        raise HTTPException(500,f'Restore failed: {e}')
    finally:
        conn.close()


def lookup_names(q_str:str=''):
    conn=db_conn()
    if not (q1(conn,'SELECT value FROM settings WHERE key=?',('lookup_enabled',)) or {}).get('value')=='1': conn.close(); raise HTTPException(403,'Disabled')
    if len(q_str)<1: conn.close(); return[]
    if USE_PG:
        r=q(conn,'SELECT id,full_name,department FROM employees WHERE full_name ILIKE ? AND status=? LIMIT 15',(f'%{q_str}%','active'))
    else:
        r=q(conn,'SELECT id,full_name,department FROM employees WHERE LOWER(full_name) LIKE ? AND status=? LIMIT 15',(f'%{q_str.lower()}%','active'))
    conn.close(); return r

@app.get('/api/lookup/employee/{eid}')
def lookup_emp(eid:int):
    conn=db_conn();today=date.today().isoformat()
    e=q1(conn,'SELECT e.full_name,e.department,e.shift_type,l.name acc_name,fp.name food_pref FROM employees e LEFT JOIN locations l ON e.accommodation_id=l.id LEFT JOIN food_preferences fp ON e.food_pref_id=fp.id WHERE e.id=? AND e.status=?',(eid,'active'))
    if not e: conn.close(); raise HTTPException(404)
    susp=q1(conn,'SELECT id FROM suspensions WHERE employee_id=? AND is_active=1 AND start_date<=? AND (end_date IS NULL OR end_date>=?)',(eid,today,today))
    absent=q1(conn,'SELECT id FROM attendance WHERE employee_id=? AND att_date=? AND status=?',(eid,today,'absent'))
    fasting=q1(conn,'SELECT id FROM fasting_records WHERE employee_id=? AND is_active=1 AND start_date<=? AND (end_date IS NULL OR end_date>=?)',(eid,today,today))
    vac=q1(conn,'SELECT id FROM vacation_records WHERE employee_id=? AND is_active=1 AND start_date<=? AND end_date>=?',(eid,today,today))
    rules=get_rules(conn);temp_ovr=get_temp_overrides(conn,today);conn.close()
    is_ab=bool(absent);is_fa=bool(fasting);fp=e['food_pref'] or ''
    acc=e['acc_name'] or 'Accommodation';eff_shift=e['shift_type']
    if eid in temp_ovr:
        ovr=temp_ovr[eid]
        if ovr.get('override_shift_type'):eff_shift=ovr['override_shift_type']
        if ovr.get('new_acc_name'):acc=ovr['new_acc_name']
    m=meal_delivery(eff_shift,rules,is_ab,is_fa,fp)
    ln_loc=acc if m['ln_to_acc'] else('Site - packed with breakfast at '+acc if m['ln_to_site'] else 'Factory')
    dn_loc='Factory' if m['dn_to_factory'] else acc
    if vac:cs='vacation'
    elif susp:cs='suspended'
    elif absent:cs='absent'
    elif fasting:cs='fasting'
    else:cs='active'
    et={'shift1':'1st Shift (7AM-3PM)','shift2':'2nd Shift (3PM-11PM)','shift3':'3rd Shift (11PM-7AM)','normal':'Normal (9AM-6PM)','site':'Site Worker'}
    return{**e,'current_status':cs,'get_bk':m['get_bk'],'get_ln':m['get_ln'],'get_dn':m['get_dn'],'bk_loc':acc,'ln_loc':ln_loc,'dn_loc':dn_loc,'iftar_kit':m['iftar'],'shift_label':et.get(eff_shift,eff_shift)}

@app.get('/api/lookup/settings')
def lk_settings(_=Depends(require('admin'))):
    conn=db_conn();s=q1(conn,'SELECT value FROM settings WHERE key=?',('lookup_enabled',));conn.close();return{'lookup_enabled':s['value'] if s else '1'}

@app.post('/api/lookup/settings')
def upd_lk(data:dict,_=Depends(require('admin'))):
    conn=db_conn();upsert_setting(conn,'lookup_enabled',data.get('lookup_enabled','1'));conn.close();return{'ok':True}

@app.post('/api/maintenance/expire')
def expire_all(_=Depends(require('admin'))):
    conn=db_conn();today=date.today().isoformat()
    expired=q(conn,'SELECT id,employee_id FROM suspensions WHERE is_active=1 AND end_date IS NOT NULL AND end_date < ?',(today,))
    for s in expired:
        exe(conn,'UPDATE suspensions SET is_active=0 WHERE id=?',(s['id'],))
        if q1(conn,'SELECT COUNT(*) c FROM suspensions WHERE employee_id=? AND is_active=1',(s['employee_id'],))['c']==0:
            exe(conn,"UPDATE employees SET status='active' WHERE id=? AND status='suspended'",(s['employee_id'],))
    orphaned=q(conn,"SELECT id FROM employees WHERE status='suspended' AND id NOT IN (SELECT employee_id FROM suspensions WHERE is_active=1)")
    for o in orphaned: exe(conn,"UPDATE employees SET status='active' WHERE id=?",(o['id'],))
    conn.close();return{'expired_suspensions':len(expired),'orphaned_fixed':len(orphaned)}

@app.get('/api/health')
def health():
    try:
        conn=db_conn();q1(conn,'SELECT 1 AS ok');conn.close()
        return{'status':'ok','database':'postgresql' if USE_PG else 'sqlite','version':'5.2'}
    except Exception as e:
        raise HTTPException(500,f'DB error: {e}')

static_dir=os.path.join(APP_DIR,'static')
NO_CACHE={'Cache-Control':'no-cache, no-store, must-revalidate','Pragma':'no-cache','Expires':'0'}
if os.path.isdir(static_dir):
    app.mount('/static',StaticFiles(directory=static_dir),name='static')

@app.get('/{full_path:path}',response_class=HTMLResponse)
def spa(full_path:str=''):
    idx=os.path.join(static_dir,'index.html')
    if os.path.exists(idx): return FileResponse(idx,headers=NO_CACHE)
    return HTMLResponse('<h1>Frontend missing</h1>')

if __name__=='__main__':
    if not USE_PG:
        print('\n⚠️  WARNING: DATABASE_URL not set. Using SQLite (data lost on redeploy).')
        print('   Set DATABASE_URL to a PostgreSQL connection string for production.\n')
    init_db()
    db_type='PostgreSQL ✅' if USE_PG else 'SQLite ⚠️  (local only)'
    print(f'{"="*54}')
    print(f'  GeometryHome WFMS v5.2  |  DB: {db_type}')
    print(f'  Listening on port {PORT}')
    print(f'{"="*54}')
    uvicorn.run('app:app',host='0.0.0.0',port=PORT,reload=False)
